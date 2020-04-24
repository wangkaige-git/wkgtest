#coding=utf-8
import openpyxl

class ReadExcel():

    def __init__(self):
        exceldir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        self.excelpath = exceldir +"\\data\\test.xlsx"

    def get_worksheet(self,sheet=0):
        workexcel = openpyxl.load_workbook(self.excelpath)
        return workexcel.worksheets[sheet]

    #封装获取某个单元格内容的函数
    def get_cell_value(self,cell_key,sheet=0):
        worksheet = self.get_worksheet(sheet)
        return worksheet[cell_key].value

    #获取单元格整行的数据
    def get_data(self,sheet=0):
        worksheet = self.get_worksheet(sheet)
        rows = worksheet.max_row
        clos = worksheet.max_column
        ex_datas = []
        for i in range(rows):
            ex_datas.append([])
        for i in range(1,rows+1):
            for j in range(1,clos+1):
                ex_datas[i-1].append(worksheet.cell(i, j).value)
        return ex_datas